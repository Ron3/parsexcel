#coding:utf-8
'''
Created on 2014-12-15

@author: Shawn
'''

import os

from openpyxl import load_workbook


class Parser(object):
    """
    解析指定路径下全部的excel文件
    """
    ''' 可以解析的 excel 文件类型 '''
    # EXCEL_FILEL_TYPE_LIST = ['xlsx', 'xlsm', 'xls']
    EXCEL_FILEL_TYPE_LIST = ['xlsx', 'xlsm']

    def __init__(self, path, initSheet=set()):
        """
        解析指定的文件
        :param path: 路径
        :param initSheet: 只解析指定的 sheetName 的表
        :return:
        """
        self.path = path

        ''' excel 文件及其内容的实例 '''
        self.excel = {}

        ''' 只需要初始化的表 '''
        self.initSheet = initSheet


    def __getitem__(self, item):
        """
        根据文件名获得Workbook实例
        :param item:
        :return:
        """
        return self.excel.get(item)


    def parse_all(self):
        """
        解析所有文件
        :return:
        """
        for fn in self.excelFilenameS:
            filepath = os.path.join(self.floder, fn)
            return filepath


    @property
    def floder(self):
        """
        文件夹路径
        :return:
        """
        return os.path.join(os.getcwd(), self.path)


    @property
    def excelFilenameS(self):
        """
        获得路径下所有excel文件的文件名
        :return:
        """
        filenameS = []
        for d, fd, fl in os.walk(self.floder):
            filenameS = fl
            break

        excelFilesnameS = []
        for filename in filenameS:
            sufix = os.path.splitext(filename)[1][1:]
            if sufix in self.EXCEL_FILEL_TYPE_LIST:
                excelFilesnameS.append(filename)

        return excelFilesnameS


    def load(self):
        """
        加载 excel 文件
        :return:
        """
        for efn in self.excelFilenameS:
            ''' excel file name '''
            ap = self.getAP(efn)
            index = ap.find("~$")
            if index >= 0:
                continue
            # wb = load_workbook(filename=ap, read_only=True)
            wb = load_workbook(filename=ap, read_only=True, data_only=True)
            self.excel[efn] = wb

        ''' 生成 infoArray 数据缓存 '''
        self.setInfoArray()


    def getAP(self, filename):
        """
        获得绝对路径
        :param filename:
        :return:
        """
        return os.path.join(self.floder, filename)


    def setInfoArray(self):
        """
        生成infoArray格式的数据缓存
        :return:
        """
        for ws in self.getAllWorksheet():
            if ws.title not in self.initSheet:
                ''' 不在解表的范围 '''
                continue

            infoArray = []
            for r in ws.get_squared_range(1, 1, ws.max_column, ws.max_row):
                infoArray.append([c.value for c in r])
            ws.infoArray = infoArray


    def getAllWorksheet(self):
        """
        获得所有的Worrsheet实例
        :return:
        """
        wsS = []
        for wb in self.excel.values():
            wsS.extend(Parser.getAllWorksheetByWokrbook(wb))

        return wsS


    @staticmethod
    def getAllWorksheetByWokrbook(wb):
        """
        获得指定工作薄的所有工作表
        :param wb: Workbook()
        :return:
        """
        wsS = []
        for wsn in wb.sheetnames:
            wsS.append(wb.get_sheet_by_name(wsn))

        return wsS